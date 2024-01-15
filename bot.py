import logging
from aiogram import executor
from aiogram import Bot, Dispatcher, types
from aiogram.contrib.middlewares.logging import LoggingMiddleware
from aiogram.types import ParseMode, InlineKeyboardMarkup, InlineKeyboardButton
from Keyboards.inline import *
from Keyboards.default import buyurtma_berish, locations, parol
from aiogram.types import ReplyKeyboardRemove
import sqlite3

son = {
    'user_id': 1
}


API_TOKEN = '6539001396:AAHTx8uhcJSf2qqBLWHIVwn-LCvkrVnObEE'

from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters.state import State, StatesGroup

logging.basicConfig(level=logging.INFO)
bot = Bot(token=API_TOKEN, parse_mode='HTML')
dp = Dispatcher(bot, storage=MemoryStorage())
dp.middleware.setup(LoggingMiddleware())



class Shogirdcha(StatesGroup):
    loc_yetkazib_berish = State()
    buyurtmachi = State()


savatchamiz_user = {
    'user_id': [],
}

conn = sqlite3.connect('stats.db')
cursor = conn.cursor()

cursor.execute('''CREATE TABLE IF NOT EXISTS stats
                  (id INTEGER PRIMARY KEY AUTOINCREMENT,
                   user_id INTEGER,
                   date DATE)''')
conn.commit()

def record_stat(user_id):
    cursor.execute("INSERT INTO stats (user_id, date) VALUES (?, DATE('now'))", (user_id,))
    conn.commit()


@dp.message_handler(commands=['stats'])
async def show_stats(message: types.Message):
    cursor.execute("SELECT COUNT(DISTINCT user_id) FROM stats")
    total_users = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(DISTINCT user_id) FROM stats WHERE date = DATE('now')")
    today_users = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM stats")
    total_requests = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM stats WHERE date = DATE('now')")
    today_requests = cursor.fetchone()[0]

    text = f"📊 Botdan foydalanish statistikasi:\n" \
           f" ├ Jami foydalanuvchilar: {total_users}\n" \
           f" ├ Bugungi foydalanuvchilar: {today_users}\n" \
           f" ├ Jami so'rovlar: {total_requests}\n" \
           f" └ Bugungi so'rovlar: {today_requests}"

    await message.reply(text)







@dp.message_handler(commands='start')
async def boshlaovchi(message: types.Message):
    son[message.from_user.id] = 1
    print(son)
    await message.answer('Buyurtmani birga joylashtiramizmi? 🤗', reply_markup=ReplyKeyboardRemove())
    await message.answer('''
Buyurtma berishni boshlash uchun 🛒 Buyurtma qilish tugmasini bosing
 
Shuningdek, aksiyalarni ko'rishingiz va bizning filiallar bilan tanishishingiz mumkin

<a href="https://telegra.ph/Taomnoma-09-30">Oqtepa Lavash Menu</a>
''', reply_markup=asosiy_menyu)
    await record_stat(message.from_user.id)


@dp.callback_query_handler(text='Buyurtma')
async def buyurtmalar(call: types.CallbackQuery):
    await call.answer('Buyurtma Bering❤️')
    await call.message.answer('Buyurtma turini tanlang', reply_markup=buyurtma_berish)


@dp.callback_query_handler(text='bizhaqqimizda')
async def biz(call: types.CallbackQuery):
    await call.message.answer('''
    
Biz O‘zbekiston bozorida 12 yildan beri faoliyat yuritamiz va bugungi kunda butun mamlakat bo‘ylab 50 dan ortiq filiallarimiz mavjud. Mazali va to‘yimli taomlar, qulay narxlar, tez yetkazib berish xizmatidan mamnun mijozlar yana va yana bizni tanlamoqda.

Qaynoqqina va mazali lavashlarimiz, shaurmayu donerlarimiz, gamburger va pitsalarimizdan albatta tatib ko'rishingizni tavsiya qilamiz va buyurtmangizga tovuq go'shtidan yangiliklarimizni qo'shishni unutmang!

Yetkazib berish xizmati:  +998781500030
Sayt (https://oqtepalavash.uz/) | Facebook (http://fb.me/oqtepalavash.official) | Instagram (https://www.instagram.com/oqtepalavash.official)   

 
''', reply_markup=ortga1)


@dp.callback_query_handler(text='ortga')
async def ortga(call: types.CallbackQuery):
    await call.message.answer('Buyurtmani birga joylashtiramizmi? 🤗')
    await call.message.answer('''
    Buyurtma berishni boshlash uchun 🛒 Buyurtma qilish tugmasini bosing

    Shuningdek, aksiyalarni ko'rishingiz va bizning filiallar bilan tanishishingiz mumkin
    ''', reply_markup=asosiy_menyu)


@dp.message_handler(text='Ortga')
async def back(message: types.Message):
    await message.answer('ortga qaytildi', reply_markup=ReplyKeyboardRemove())
    await message.answer('''
    Buyurtma berishni boshlash uchun 🛒 Buyurtma qilish tugmasini bosing

    Shuningdek, aksiyalarni ko'rishingiz va bizning filiallar bilan tanishishingiz mumkin
    ''', reply_markup=asosiy_menyu)



@dp.message_handler(text='⬅️Ortga')
async def exit(message: types.Message):
    await message.answer('Buyurtmani birga joylashtiramizmi? 🤗')
    await message.answer('''
Buyurtma berishni boshlash uchun 🛒 Buyurtma qilish tugmasini bosing
 
Shuningdek, aksiyalarni ko'rishingiz va bizning filiallar bilan tanishishingiz mumkin

<a href="https://telegra.ph/Taomnoma-09-30">Oqtepa Lavash Menu</a>
        ''', reply_markup=asosiy_menyu, )


@dp.message_handler(text='Eltib berish🛵')
async def loc_user(message: types.Message):
    await message.answer("<b>Eltib berish</b> uchun <b>geo-joylashuvni</b> jo'nating yoki manzilni tanlang",reply_markup=locations)
    await Shogirdcha.loc_yetkazib_berish.set()


@dp.message_handler(text='⬅️ Ortga')
async def exit1(message:types.Message):
    await message .answer('Buyurtmani birga joylashtiramizmi? 🤗',reply_markup=asosiy_menyu)




@dp.message_handler(content_types=types.ContentType.LOCATION, state=Shogirdcha.loc_yetkazib_berish)
async def location_saver(message: types.Message, state: FSMContext):
    global LL
    LL = message.location




    await message.answer('Locatsiya qabul qilindi', reply_markup=ReplyKeyboardRemove())
    photo = open('pictures/kategoriya.jpg', 'rb')
    await message.answer('Kategoriyalrdan birini tanlang')
    await message.answer_photo(photo=photo)
    await message.answer("⏬---------------------------------⏬", reply_markup=snakes)
    await state.finish()


@dp.callback_query_handler(text='tovuq')
async def tovuqlar(call: types.CallbackQuery):
    photo = open('pictures/qarsildoq.jpg', 'rb')
    await call.message.answer_photo(photo=photo, )
    await call.message.answer("🍗Qarsildoq Jojalar", reply_markup=qarsildoq_oyoqlar)


@dp.callback_query_handler(text='box')
async def box(call: types.CallbackQuery):
    photo = open('pictures/JOJABOX.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('''
    Jo'ja box
Narxi:   26 000 so'm
Tavsif: Strips 3 dona, kartoshka fri o'rta va ketchup
Miqdorini tanlang yoki kiriting
    ''', reply_markup=savat)


@dp.callback_query_handler(text='stips')
async def stips(call: types.CallbackQuery):
    photo = open('pictures/stipsdona.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('''
    Stips 5 dona
Narxi:   26 000 so'm
Tavsif: Qarsildoq panirovkadagi uzun shaklda kesilgan tovuq bo'laklari
Miqdorini tanlang yoki kiriting
    ''', reply_markup=savat)


@dp.callback_query_handler(text='stip')
async def stips(call: types.CallbackQuery):
    photo = open('pictures/Stips3don.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('''
    Stips 3 dona
Narxi:   16 000 so'm
Tavsif: Qarsildoq panirovkadagi uzun shaklda kesilgan tovuq bo'laklari
Miqdorini tanlang yoki kiriting
    ''', reply_markup=savat)


@dp.callback_query_handler(text='bayt')
async def stips(call: types.CallbackQuery):
    photo = open('pictures/bayts.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('''
    Bayts
Narxi:   16 000 so'm
Tavsif: Qarsildoq panirovkadagi tovuq bo'laklari 
Miqdorini tanlang yoki kiriting
    ''', reply_markup=savat)


@dp.callback_query_handler(text='salat')
async def salat(call: types.CallbackQuery):
    photo = open('pictures/importtant_salat.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer("🥗 Salatlar", reply_markup=salats)


@dp.callback_query_handler(text='kapriz')
async def kap(call: types.CallbackQuery):
    photo = open('pictures/kapriz.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('''
    Mujskoy kapriz
Narxi:   25 000 so'm
Tavsif: Dudlangan kolbasa, kurka, qazi, pishloq, mayonez
Miqdorini tanlang yoki kiriting
    ''', reply_markup=savat_salatlar)


@dp.callback_query_handler(text='sezat')
async def kap(call: types.CallbackQuery):
    photo = open('pictures/sezar.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('''
    Sezar
Narxi:   23 000 so'm
Tavsif: Tovuq filesi, pomidor, aysberg, pishloq, kruton, sarimsoq sousi.
Miqdorini tanlang yoki kiriting
    ''', reply_markup=savat_salatlar)


@dp.callback_query_handler(text='salat_exit')
async def jojaexit(call: types.CallbackQuery):
    photo = open('pictures/importtant_salat.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer("🥗Salatlar", reply_markup=salats)


@dp.callback_query_handler(text='exits')
async def exits(call: types.CallbackQuery):
    photo = open('pictures/2023-08-10 14.15.12.jpg', 'rb')
    await call.message.answer_photo(photo=photo, reply_markup=snakes)


@dp.callback_query_handler(text='qarsildoq_exit')
async def qarsildoqexit(call: types.CallbackQuery):
    photo = open('pictures/kategoriya.jpg', 'rb')
    await call.message.answer_photo(photo=photo, reply_markup=snakes)


@dp.callback_query_handler(text='joja_exit')
async def jojaexit(call: types.CallbackQuery):
    photo = open('pictures/qarsildoq.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer("🍗Qarsildoq Jo'jalar", reply_markup=qarsildoq_oyoqlar)


@dp.callback_query_handler(text='ichimlik_cola')
async def ichimlik(call: types.CallbackQuery):
    photo = open('pictures/important.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer("🥤Ichimliklar", reply_markup=ichimliklar_1)


@dp.callback_query_handler(text='qaynoq_ichimlik')
async def qaynoq(call: types.CallbackQuery):
    photo = open('pictures/important1.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer("🥤Ichimliklar> Qaynoq ichimliklar", reply_markup=hot_tea)


@dp.callback_query_handler(text='qora')
async def qorachoy(call: types.CallbackQuery):
    photo = open('pictures/qora.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('''
    Qora choy
Narxi:   3 000 so'm
Tavsif: Qora choy
Miqdorini tanlang yoki kiriting''', reply_markup=choylar)


@dp.callback_query_handler(text='kok')
async def kokchoy(call: types.CallbackQuery):
    photo = open('pictures/2023-08-10 14.15.12.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('''
    Ko'k choy
Narxi:   3 000 so'm
Tavsif: Ko'k choy
Miqdorini tanlang yoki kiriting
    ''', reply_markup=choylar)


@dp.callback_query_handler(text='kok2')
async def limonchoy(call: types.CallbackQuery):
    photo = open('pictures/limon.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('''
    Ko'k choy
Narxi:   3 000 so'm
Tavsif: Ko'k choy
Miqdorini tanlang yoki kiriting
    ''', reply_markup=choylar)


@dp.callback_query_handler(text='hot_exit')
async def hotexit(call: types.CallbackQuery):
    photo = open('pictures/important1.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer("🥤Ichimliklar > Qaynoq ichimliklar", reply_markup=ichimliklar_1)


@dp.callback_query_handler(text='choylar_exit')
async def ortga(call: types.CallbackQuery):
    photo = open('pictures/important.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer("🥤Ichimliklar > Qaynoq ichimliklar", reply_markup=hot_tea)


@dp.callback_query_handler(text='yaxna_ichimlik')
async def yaxna(call: types.CallbackQuery):
    photo = open('pictures/yaxna ichimliklar.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer("🥤Ichimliklar > Yahna ichimliklar", reply_markup=yaxna_water)


@dp.callback_query_handler(text='pepsi')
async def pepsi(call: types.CallbackQuery):
    photo = open('pictures/pepsi.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer("🥤Ichimliklar > Yahna ichimliklar > pepsi", reply_markup=pepsi_water)


@dp.callback_query_handler(text='ichimlik_exit')
async def pepsi(call: types.CallbackQuery):
    photo = open('pictures/kategoriya.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer("Kategoriyalardan birini tanlang", reply_markup=snakes)


@dp.callback_query_handler(text='pepsi_exit')
async def pepsiexit(call: types.CallbackQuery):
    photo = open('pictures/yaxna ichimliklar.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer("🥤Ichimliklar > Yahna ichimliklar", reply_markup=yaxna_water)


@dp.callback_query_handler(text='pepsi1_exit')
async def pepsiexit(call: types.CallbackQuery):
    photo = open('pictures/yaxna ichimliklar.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer("🥤Ichimliklar > Yahna ichimliklar", reply_markup=pepsi_water)


@dp.callback_query_handler(text='pepsi1.5')
async def pepsi(call: types.CallbackQuery):
    photo = open('pictures/pepsi1.5.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('''
    Pepsi 1.5
Narxi:   17 000 so'm
Tavsif: Pepsi 1.5
Miqdorini tanlang yoki kiriting
    ''', reply_markup=pepsi_savat)







@dp.callback_query_handler(text='pepsi0.4')
async def pepsi(call: types.CallbackQuery):
    photo = open('pictures/pepsi0.4.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('''
Pepsi 0.4
Narxi:   8 000 so'm
Tavsif: Pepsi 0.4
Miqdorini tanlang yoki kiriting
    ''', reply_markup=pepsi_savat)


@dp.callback_query_handler(text='pepsi0.3')
async def pepsi(call: types.CallbackQuery):
    photo = open('pictures/pepsi03.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('''
Pepsi 0.3 L
Narxi:   6 000 so'm
Tavsif: Pepsi 0.3 L
Miqdorini tanlang yoki kiriting
    ''', reply_markup=pepsi_savat)


@dp.callback_query_handler(text='dolina_exit')
async def dolinaexit(call: types.CallbackQuery):
    photo = open('pictures/pepsi.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer("🥤Ichimliklar > Yahna ichimliklar", reply_markup=yaxna_water)


@dp.callback_query_handler(text='pepsi0.5')
async def pepsi(call: types.CallbackQuery):
    photo = open('pictures/pepsi05.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('''
Pepsi 0.5
Narxi:   9 000 so'm
Tavsif: Pepsi 0.5
Miqdorini tanlang yoki kiriting
        ''', reply_markup=pepsi_savat)


@dp.callback_query_handler(text='dolina')
async def pepsi(call: types.CallbackQuery):
    photo = open('pictures/dolina.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('''
Sochnaya dolina 1L
Narxi:   16 000 so'm
Tavsif: Sochnaya dolina 1L
Miqdorini tanlang yoki kiriting
        ''', reply_markup=dolina_choy)


@dp.callback_query_handler(text='mirinda')
async def mirinda(call: types.CallbackQuery):
    photo = open('pictures/mirinda.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('''
    Mirinda 0.4L
Narxi:   8 000 so'm
Tavsif: Mirinda 0.4L, 
Miqdorini tanlang yoki kiriting
    ''', reply_markup=mirinda1)


@dp.callback_query_handler(text='mirinda1')
async def mirinda(call: types.CallbackQuery):
    photo = open('pictures/yaxna ichimliklar.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer("🥤Ichimliklar > Yahna ichimliklar", reply_markup=yaxna_water)


@dp.callback_query_handler(text='mountain')
async def mirinda(call: types.CallbackQuery):
    photo = open('pictures/mountindew.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('''
    Mountain Dew 0.4L
Narxi:   8 000 so'm
Tavsif: Mountain Dew 0.4L, 
Miqdorini tanlang yoki kiriting
    ''', reply_markup=mountain)


@dp.callback_query_handler(text='mountain_exit')
async def mirinda(call: types.CallbackQuery):
    photo = open('pictures/yaxna ichimliklar.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer("🥤Ichimliklar > Yahna ichimliklar", reply_markup=yaxna_water)


@dp.callback_query_handler(text='icetea')
async def hello(call: types.CallbackQuery):
    photo = open('pictures/icetea.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('''
    Ice tea 0.5
Narxi:   15 000 so'm
Tavsif: Ice tea 0.5
Miqdorini tanlang yoki kiriting
    ''', reply_markup=icetea)


@dp.callback_query_handler(text='icetea_exit')
async def hello(call: types.CallbackQuery):
    photo = open('pictures/yaxna ichimliklar.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('🥤Ichimliklar > Yahna ichimliklar', reply_markup=yaxna_water)


@dp.callback_query_handler(text='yaxna_exit')
async def hello(call: types.CallbackQuery):
    photo = open('pictures/important.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('🥤Ichimliklar', reply_markup=ichimliklar_1)


@dp.callback_query_handler(text='sous')
async def hello(call: types.CallbackQuery):
    photo = open('pictures/first_sous.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('🍅Souslar', reply_markup=souslar)


@dp.callback_query_handler(text='ketchup')
async def hello(call: types.CallbackQuery):
    photo = open('pictures/sous1.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('''
    <b>Ketchup</b>
Narxi:  <b> 3 000 so'm</b>
Tavsif: Ketchup
Miqdorini tanlang yoki kiriting
    ''', reply_markup=sous1)


@dp.callback_query_handler(text='sous1_exit')
async def sous(call: types.CallbackQuery):
    photo = open('pictures/first_sous.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('🍅Souslar', reply_markup=souslar)


@dp.callback_query_handler(text='chili')
async def sous(call: types.CallbackQuery):
    photo = open('pictures/chili.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('''
    <b>Chili sous</b>
Narxi:    <b>3 000 so'm</b>
Tavsif: Chili sous
Miqdorini tanlang yoki kiriting
    ''', reply_markup=sous1)


@dp.callback_query_handler(text='pishloq_sous')
async def sous(call: types.CallbackQuery):
    photo = open('pictures/pishloqli.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('''
<b>Pishloqli  sous</b>
Narxi:    <b>3 000 so'm</b>
Tavsif: Pishloqli  sous
Miqdorini tanlang yoki kiriting
    ''', reply_markup=sous1)


@dp.callback_query_handler(text='oq_sous')
async def sous(call: types.CallbackQuery):
    photo = open('pictures/oqsous.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('''
<b>Oq Sous</b>
Narxi:   <b> 3 000 so'm</b>
Tavsif: Oq Sous
Miqdorini tanlang yoki kiriting
    ''', reply_markup=sous1)


@dp.callback_query_handler(text='lavash')
async def lavash(call: types.CallbackQuery):
    photo = open('pictures/lavash.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('🌯Lavashlar', reply_markup=lavash_button)


@dp.callback_query_handler(text='lavash_exit')
async def lavash(call: types.CallbackQuery):
    photo = open('pictures/kategoriya.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('Kategoriyalardan birini tanlang', reply_markup=snakes)


@dp.callback_query_handler(text='savat_exit')
async def savat_exit(call: types.CallbackQuery):
    photo = open('pictures/lavash.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('🌯Lavashlar', reply_markup=lavash_button)


@dp.callback_query_handler(text='original')
async def original(call: types.CallbackQuery):
    photo = open('pictures/original.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('''
    Original lavash
Narxi:   28 000 so'm
Tavsif: Yupqa lavash non, pomidor, chips, mol ,go'shti\nqizil sous, mayonez.
Miqdorini tanlang yoki kiriting
    ''', reply_markup=lavash_savat)


@dp.callback_query_handler(text='org_kichik')
async def original(call: types.CallbackQuery):
    photo = open('pictures/orgkichik.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('''
Original kichik lavash
Narxi:   23 000 so'm
Tavsif: Yupqa lavash non, pomidor, chips, mol go'shti, qizil\nsous, mayonez, pishloq.
Miqdorini tanlang yoki kiriting
    ''', reply_markup=lavash_savat)


@dp.callback_query_handler(text='pishloq')
async def original(call: types.CallbackQuery):
    photo = open('pictures/pishloqli.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('''
Pishloqli lavash
Narxi:   31 000 so'm
Tavsif: Yupqa lavash non, pomidor, chips, mol go'shti, qizil\nsous, mayonez, pishloq.
Miqdorini tanlang yoki kiriting
    ''', reply_markup=lavash_savat)


@dp.callback_query_handler(text='pish_kichik')
async def original(call: types.CallbackQuery):
    photo = open('pictures/pishloqlikich.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('''Pishloqli kichik lavash
Narxi:   26 000 so'm
Tavsif: Yupqa lavash non, pomidor, chips, mol go'shti, qizil\nsous, mayonez, pishloq.
Miqdorini tanlang yoki kiriting
    ''', reply_markup=lavash_savat)


@dp.callback_query_handler(text='tandir')
async def original(call: types.CallbackQuery):
    photo = open('pictures/TANDIR.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('''
        Tandir lavash
Narxi:   29 000 so'm
Tavsif: Tandir pechida pishirilgan yupqa lavash non, pomidor,\nchips, mol go'shti, qizil sous, mayonez, kunjut.
Miqdorini tanlang yoki kiriting
        ''', reply_markup=lavash_savat)


@dp.callback_query_handler(text='pish_tandir')
async def original(call: types.CallbackQuery):
    photo = open('pictures/pishtandir.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('''
Pishloqli tandir lavash
Narxi:   32 000 so'm
Tavsif: Tandirda pishirilgan yupqa lavash non, pomidor, chips,\nmol go'shti, pishloq, qizil sous, mayonez, kunjut.
Miqdorini tanlang yoki kiriting
        ''', reply_markup=lavash_savat)


@dp.callback_query_handler(text='set')
async def set(call: types.CallbackQuery):
    photo = open('pictures/setlar.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('🍟🌯🥤Setlar', reply_markup=setlar)


@dp.callback_query_handler(text='setlar_exit')
async def set(call: types.CallbackQuery):
    photo = open('pictures/kategoriya.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('Kategoriyalardan birini tanlang', reply_markup=snakes)


@dp.callback_query_handler(text='oqtepa')
async def set(call: types.CallbackQuery):
    photo = open('pictures/oqtepa.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('''
    Oqtepa seti
Narxi:   17 000 so'm
Tavsif: Kartoshka fri o'rta, Pepsi 0.4L, 
Miqdorini tanlang yoki kiriting
    ''', reply_markup=set_savat)


@dp.callback_query_handler(text='juftlik')
async def set(call: types.CallbackQuery):
    photo = open('pictures/juftlik.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('''
"Juftlik" Seti
Narxi:   58 000 so'm
Tavsif: Klab sendvich, strips 3 dona, Pepsi 0.4L\n2 dona, ketchup 2 dona
Miqdorini tanlang yoki kiriting
    ''', reply_markup=set_savat)


@dp.callback_query_handler(text='baraka')
async def set(call: types.CallbackQuery):
    photo = open('pictures/baraka.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('''
"Baraka" Seti
Narxi:   134 000 so'm
Tavsif: Assorti pitsa, kartoshka fri o'rta 3 dona,\nPepsi 0.4L 3 dona, ketchup 3 dona
Miqdorini tanlang yoki kiriting
    ''', reply_markup=set_savat)


@dp.callback_query_handler(text='set_exit')
async def set(call: types.CallbackQuery):
    photo = open('pictures/setlar.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('🍟🌯🥤Setlar', reply_markup=setlar)


@dp.callback_query_handler(text='pitsa')
async def set(call: types.CallbackQuery):
    photo = open('pictures/pitsalar.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('🍕Pitsalar', reply_markup=pitsa)


@dp.callback_query_handler(text='assorti')
async def set(call: types.CallbackQuery):
    photo = open('pictures/assorti.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('''
        Assorti pitsa
Narxi:   89 000 so'm
Tavsif: Oq sous, zaytun, qo'ziqorin, bulg'or\nqalampiri, pomidor, dudlangan kurka,\ndudlangan kolbasa, mol go'shti, sosiska,\nMozzarella va Akbel pishloqlari.
Miqdorini tanlang yoki kiriting
        ''', reply_markup=pitsa_savat)


@dp.callback_query_handler(text='peperonni')
async def set(call: types.CallbackQuery):
    photo = open('pictures/peperoni.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('''
        Pepperoni pitsa
Narxi:   75 000 so'm
Tavsif: “OQTEPA” pomidor sousi, dudlangan\nkolbasa,Mozzarella va Akbel pishlog'i 
Miqdorini tanlang yoki kiriting
                ''', reply_markup=pitsa_savat)


@dp.callback_query_handler(text='goshtli')
async def set(call: types.CallbackQuery):
    photo = open('pictures/goshtli.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('''
        Go'shtli pitsa
Narxi:   87 000 so'm
Tavsif: Tomato sauce “OQTEPA”, chicken meat,\nbell pepper, beef, tomatoes, Mozzarella and Akbel cheese
Miqdorini tanlang yoki kiriting
                ''', reply_markup=pitsa_savat)


@dp.callback_query_handler(text='qazi')
async def set(call: types.CallbackQuery):
    photo = open('pictures/qazi.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('''
        Qazi pizza
Narxi:   90 000 so'm
Tavsif: “OQTEPA” pomidor sousi, "Brunswick"\nshirin piyoz halqalari, qazi, Mozzarella va Akbel pishloqlari
Miqdorini tanlang yoki kiriting
                ''', reply_markup=pitsa_savat)


@dp.callback_query_handler(text='tovuqli_pitsa')
async def set(call: types.CallbackQuery):
    photo = open('pictures/tovuqli_pitsa.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('''
        Tovuqli pitsa
Narxi:   75 000 so'm
Tavsif: OQTEPA pomidor sousi, kurka, tovuq,\nqo'ziqorin, zaytun, pishloq, oregano
Miqdorini tanlang yoki kiriting
                ''', reply_markup=pitsa_savat)


@dp.callback_query_handler(text='pitsa_Exit')
async def set(call: types.CallbackQuery):
    photo = open('pictures/setlar.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('Kategorilardan birini tanlang', reply_markup=snakes)


@dp.callback_query_handler(text='pitsa_exit')
async def set(call: types.CallbackQuery):
    photo = open('pictures/pitsalar.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('🍕Pitsalar', reply_markup=pitsa)


@dp.callback_query_handler(text='burger')
async def set(call: types.CallbackQuery):
    photo = open('pictures/burgerlar.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('🍔 Burger va donerlar', reply_markup=burgerlar)


@dp.callback_query_handler(text='gamburger')
async def set(call: types.CallbackQuery):
    photo = open('pictures/gamburger.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('''
    Gamburger
Narxi:   22 000 so'm
Tavsif: Shirin bulochka, maxsus sous, aysberg,\ntuzlangan bodring, mol go'shtidan kotlet,\npomidor, "Brunswick" shirin piyoz halqalari
Miqdorini tanlang yoki kiriting
    ''', reply_markup=burgers_savat)


@dp.callback_query_handler(text='chizburger')
async def set(call: types.CallbackQuery):
    photo = open('pictures/chizburger.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('''
Chizburger
Narxi:   24 000 so'm
Tavsif: Shirin bulochka, maxsus sous, aysberg,\ntuzlangan bodring, mol go'shtidan kotlet,\npomidor, pishloq, "Brunswik" shirin piyoz\nhalqalari
Miqdorini tanlang yoki kiriting
    ''', reply_markup=burgers_savat)


@dp.callback_query_handler(text='big_burger')
async def set(call: types.CallbackQuery):
    photo = open('pictures/big_burger.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('''
Big burger
Narxi:   33 000 so'm
Tavsif: Shirin bulochka, maxsus sous, aysberg,\ntuzlangan bodring, ikkita mol go'shtidan kotlet,\npomidor, "Brunswick" shirin piyoz halqalari
Miqdorini tanlang yoki kiriting
    ''', reply_markup=burgers_savat)


@dp.callback_query_handler(text='bigchizburger')
async def set(call: types.CallbackQuery):
    photo = open('pictures/bigchiz.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('''
Big Chizburger
Narxi:   37 000 so'm
Tavsif: Shirin bulochka, maxsus sous, aysberg,\ntuzlangan bodring, ikkita mol go'shtidan kotlet,\npomidor, pishloq, "Brunswick" shirin piyoz\nhalqalari
Miqdorini tanlang yoki kiriting
    ''', reply_markup=burgers_savat)


@dp.callback_query_handler(text='bigdoner')
async def set(call: types.CallbackQuery):
    photo = open('pictures/bigdoner.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('''
Big doner
Narxi:   26 000 so'm
Tavsif: Ekmek mualliflik noni, oq va qizil souslar,\nchiplar, mol go'shti, bodring, pomidor.
Miqdorini tanlang yoki kiriting
    ''', reply_markup=burgers_savat)


@dp.callback_query_handler(text='shaurma')
async def set(call: types.CallbackQuery):
    photo = open('pictures/shaurma.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('''
Shaurma
Narxi:   22 000 so'm
Tavsif: Tandirli pita noni, mol go'shti, bodring,\npomidor, qizil sous, "Brunswick" shirin piyoz\nhalqalari
Miqdorini tanlang yoki kiriting
    ''', reply_markup=burgers_savat)


@dp.callback_query_handler(text='burgers_exit')
async def set(call: types.CallbackQuery):
    photo = open('pictures/kategoriya.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('Kategoriyalardan birini tanlang', reply_markup=snakes)


@dp.callback_query_handler(text='chizburger_exit')
async def set(call: types.CallbackQuery):
    photo = open('pictures/burgerlar.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('🍔 Burger va donerlar', reply_markup=burgerlar)


@dp.callback_query_handler(text='hotdog')
async def set(call: types.CallbackQuery):
    photo = open('pictures/hotdoglar.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('🌭Hot doglar', reply_markup=burgerlar)


@dp.callback_query_handler(text='haggi')
async def set(call: types.CallbackQuery):
    photo = open('pictures/haggi.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('''
    Xaggi
Narxi:   31 000 so'm
Tavsif: Baget noni, mayonez, mol go'shti, salat\nbargi, bodring, pomidor, pishloq, qizil sous,\n"Brunswick" shirin piyoz halqalari
Miqdorini tanlang yoki kiriting
    ''', reply_markup=shaurma)


@dp.callback_query_handler(text='shaurma_exit')
async def set(call: types.CallbackQuery):
    photo = open('pictures/kategoriya.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('Kategoriyalardan birini tanlang', reply_markup=snakes)


@dp.callback_query_handler(text='klab')
async def set(call: types.CallbackQuery):
    photo = open('pictures/klabsendvich.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('''
Klab sendvich
Narxi:   30 000 so'm
Tavsif: Toster non, maxsus sous, bodring,\npomidor, tovuq filesi, salat bargi, pishloq,\nkartoshka fri
Miqdorini tanlang yoki kiriting
    ''', reply_markup=klab_sendvich)


@dp.callback_query_handler(text='klab_exit')
async def set(call: types.CallbackQuery):
    photo = open('pictures/kategoriya.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('Kategoriyalardan birini tanlang', reply_markup=snakes)


@dp.callback_query_handler(text='snek')
async def set(call: types.CallbackQuery):
    photo = open('pictures/sneklar.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('🍟Sneklar', reply_markup=sneklar)


@dp.callback_query_handler(text='fri')
async def set(call: types.CallbackQuery):
    photo = open('pictures/friorta.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('''
Kartoshka fri o'rta
Narxi:   14 000 so'm
Tavsif: Kartoshka fri o'rta
Miqdorini tanlang yoki kiriting
    ''', reply_markup=sneks)


@dp.callback_query_handler(text='fri_katta')
async def set(call: types.CallbackQuery):
    photo = open('pictures/frikatta.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('''
Kartoshka fri katta
Narxi:   18 000 so'm
Tavsif: Kartoshka fri katta
Miqdorini tanlang yoki kiriting
    ''', reply_markup=sneks)


@dp.callback_query_handler(text='fri_kichik')
async def set(call: types.CallbackQuery):
    photo = open('pictures/frikichik.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('''
Kartoshka fri kichik
Narxi:   8 000 so'm
Tavsif: Kartoshka fri kichik
Miqdorini tanlang yoki kiriting
    ''', reply_markup=sneks)


@dp.callback_query_handler(text='jaydari')
async def set(call: types.CallbackQuery):
    photo = open('pictures/jaydari.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('''
Jaydari kartoshka
Narxi:   15 000 so'm
Tavsif: Jaydari kartoshka
Miqdorini tanlang yoki kiriting
    ''', reply_markup=sneks)


@dp.callback_query_handler(text='non')
async def set(call: types.CallbackQuery):
    photo = open('pictures/non.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('''
Non
Narxi:   3 000 so'm
Tavsif: Non
Miqdorini tanlang yoki kiriting
    ''', reply_markup=sneks)


@dp.callback_query_handler(text='xala')
async def set(call: types.CallbackQuery):
    global son
    sneks = InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton('➖', callback_data='minus_xlapeniyo'),
                InlineKeyboardButton(f"{son[call.message.chat.id]}", callback_data='son'),
                InlineKeyboardButton('➕', callback_data='plus_xlapeniyo')
            ],
            [
                InlineKeyboardButton("Savatga qo'shish", callback_data='savatcha_snek'),
            ],
            [
                InlineKeyboardButton("⬅Ortga", callback_data='sneks_exit'),
            ],
        ],
    )

    photo = open('pictures/xala.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('''
Xalapenyo
Narxi:   3 000 so'm
Tavsif: Xalapenyo
Miqdorini tanlang yoki kiriting
    ''', reply_markup=sneks)


@dp.callback_query_handler(text='plus_xlapeniyo')
async def xla_plus(call: types.CallbackQuery):
    global son
    user_id = str(call.message.chat.id)
    print(user_id)

    # Retrieve the user's value from the dictionary, default to 0 if not found
    print(son)
    fake_son = son.get(user_id, 0)
    fake_son += 1

    # Update the dictionary with the new value
    son[user_id] = fake_son

    # Update the existing inline keyboard with the new value
    await update_sneks_buttons(call.message.chat.id, call.message.message_id, fake_son)
    # photo = open('pictures/xala.jpg', 'rb')
    # await call.message.answer_photo(photo=photo)
    # await call.message.answer('''
    # Xalapenyo
    # Narxi:   3 000 so'm
    # Tavsif: Xalapenyo
    # Miqdorini tanlang yoki kiriting
    #     ''', reply_markup=sneks)


async def update_sneks_buttons(chat_id, message_id, new_son):
    new_buttons = InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton('➖', callback_data='minus_xlapeniyo'),
                InlineKeyboardButton(f"{new_son}", callback_data='son'),
                InlineKeyboardButton('➕', callback_data='plus_xlapeniyo')
            ],
            [
                InlineKeyboardButton("Savatga qo'shish", callback_data='savatcha_snek'),
            ],
            [
                InlineKeyboardButton("⬅Ortga", callback_data='sneks_exit'),
            ],
        ],
    )

    # Edit the existing message to update the inline keyboard
    await bot.edit_message_reply_markup(chat_id=chat_id, message_id=message_id, reply_markup=new_buttons)


@dp.callback_query_handler(text='minus_xlapeniyo')
async def minuser(call: types.CallbackQuery):
    global son
    fake_son = son[f'{call.message.chat.id}']
    fake_son -= 1
    son[f'{call.message.chat.id}'] = fake_son
    await update_sneks_buttons1(call.message.chat.id, call.message.message_id, fake_son)


async def update_sneks_buttons1(chat_id, message_id, new_son):
    new_buttons1 = InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton('➖', callback_data='minus_xlapeniyo'),
                InlineKeyboardButton(f"{new_son}", callback_data='son'),
                InlineKeyboardButton('➕', callback_data='plus_xlapeniyo')
            ],
            [
                InlineKeyboardButton("Savatga qo'shish", callback_data='savatcha_snek'),
            ],
            [
                InlineKeyboardButton("⬅Ortga", callback_data='sneks_exit'),
            ],
        ],
    )

    # Edit the existing message to update the inline keyboard
    await bot.edit_message_reply_markup(chat_id=chat_id, message_id=message_id, reply_markup=new_buttons1)


@dp.callback_query_handler(text='sneks_exit')
async def set(call: types.CallbackQuery):
    photo = open('pictures/sneklar.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('🍟Sneklar', reply_markup=sneklar)


@dp.callback_query_handler(text='snek_exit')
async def set(call: types.CallbackQuery):
    photo = open('pictures/kategoriya.jpg', 'rb')
    await call.message.answer_photo(photo=photo)
    await call.message.answer('Kategoriyalardan birini tanlang', reply_markup=snakes)


@dp.callback_query_handler(text='menu')
async def set(call: types.CallbackQuery):
    await call.message.answer('Buyurtmani birga joylashtiramizmi? 🤗', reply_markup=asosiy_menyu)


# ------------------------ADMIN------------------------


import openpyxl

# import A1 case in password.xlsx
wb = openpyxl.load_workbook('password.xlsx')

sheet = wb.active
password = sheet['A1'].value


class ADMIN(StatesGroup):
    parolcha = State()
    parol_change = State()
    check = State()
    chnage = State()


@dp.message_handler(state=ADMIN.chnage)
async def changer(message: types.Message, state=FSMContext):
    password = message.text
    workbook_saver = openpyxl.Workbook()
    sheet_saver = workbook_saver.active
    sheet_saver['A1'] = password
    workbook_saver.save('password.xlsx')
    await message.answer('parol o`zgartirildi')
    await state.finish()


@dp.message_handler(commands='admin')
async def admin(message: types.Message):
    await message.answer('Parolni kiriting !')
    await ADMIN.parolcha.set()


@dp.message_handler(state=ADMIN.parolcha, content_types=types.ContentType.TEXT)
async def check_password(message: types.Message, state=FSMContext):
    wb = openpyxl.load_workbook('password.xlsx')
    sheet = wb.active
    password = sheet['A1'].value
    if message.text == password:
        await message.answer('siz admin paneldasiz !', reply_markup=parol)
        await state.finish()
        await ADMIN.parol_change.set()

    else:
        await message.reply('Parol xato qayta urinib ko`ring !')


@dp.message_handler(text='Parol ozgartirish', state=ADMIN.parol_change)
async def change_pass(message: types.Message, state=FSMContext):
    await message.answer('Eski <b>parol</b>ni kiriting !')
    await state.finish()
    await ADMIN.check.set()


@dp.message_handler(state=ADMIN.check)
async def check_password_for_change(message: types.Message, state=FSMContext):
    if message.text == password:
        await message.answer('Yangi <b>parol</b> kiriting !')
        await state.finish()
        await ADMIN.chnage.set()
    else:
        await message.reply('Parol Xato')

from Keyboards.inline import check_oshpaz
@dp.message_handler(text='Buyurtmani tasdiqlash', state=Shogirdcha.buyurtmachi)
async def apply(message: types.Message, state=FSMContext):
    global userid
    await message.answer('Sizning Buyurtmangiz Qabul qilindi!\n\nYaqin orada javobini olasiz.')
    await message.answer('⏳')
    txt = ''
    for i in range(len(savatchamiz_user[message.from_user.id])):
        txt += f'{i + 1}.' + ' 🍟mahsulot ' + savatchamiz_user[message.from_user.id][i] + '\n'
    await bot.send_message(6498877955, txt, reply_markup=check_oshpaz)
    latitudee = LL['latitude']
    longitudeee = LL['longitude']
    await bot.send_location(6498877955, latitudee, longitudeee)
    userid = message.from_user.id




@dp.callback_query_handler(text='oshpaz_true',state=Shogirdcha.buyurtmachi)
async def oshaptrue(call: types.CallbackQuery):
    await bot.send_message( userid , "Buyurtmangiz yarim saot ichida yetkazib beriladi",reply_markup=ReplyKeyboardRemove())
    await state.finish()


@dp.callback_query_handler(text='oshpaz_false',state=Shogirdcha.buyurtmachi)
async def oshapfalse(call: types.CallbackQuery):
    await bot.send_message(userid,"Buyurtmangiz bekor qilindi",reply_markup=ReplyKeyboardRemove())
    await state.finish()








if __name__ == '__main__':
    from savatcha import dp
    executor.start_polling(dp, skip_updates=True)





